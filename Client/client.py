import os
import requests
from tkinter import Tk, filedialog, StringVar, IntVar, Scrollbar, Canvas, VERTICAL
import openpyxl
import ttkbootstrap as ttk
import threading

def process_audio_file(file_path, server_url):
    process_audio_url = f"{server_url}/process_audio"
    with open(file_path, 'rb') as f:
        response = requests.post(process_audio_url, files={'file': f})
    if response.status_code == 200:
        return response.json().get("text", "")
    print(f"Error processing {file_path}: {response.json().get('error', 'Unknown error')}")
    return None

def check_server_connection(server_url):
    try:
        response = requests.get(f"{server_url}/check_connection")
        return response.status_code == 200
    except requests.exceptions.RequestException:
        return False

class Application:
    def __init__(self, master):
        self.master = master
        self.master.title("音频资源检测")
        self.init_variables()
        self.create_widgets()
        self.load_previous_selection()
        self.set_initial_window_size()
        self.center_window()

    def init_variables(self):
        self.file_path_var = StringVar(value="选择一个XLSX文件")
        self.sheet_var = StringVar(value="选择一个Sheet")
        self.selected_headers_var = StringVar()
        self.content_header_var = StringVar(value="选择原文本的表头")
        self.folder_path_var = StringVar(value="选择一个文件夹")
        self.save_folder_path_var = StringVar(value="选择结果保存文件夹")
        self.server_url_var = StringVar(value="http://localhost:5000")
        self.selected_headers = []
        self.content_header = "文本"
        self.content_header_var.trace_add('write', self.save_content_header_selection)

    def create_widgets(self):
        self.create_file_selection_widgets()
        self.create_sheet_selection_widgets()
        self.create_header_selection_widgets()
        self.create_content_header_widgets()
        self.create_folder_selection_widgets()
        self.create_server_widgets()
        self.create_action_buttons()

    def create_file_selection_widgets(self):
        ttk.Button(self.master, text="选择XLSX文件", command=self.select_file, bootstyle="outline-primary").grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        ttk.Label(self.master, textvariable=self.file_path_var, bootstyle="info").grid(row=0, column=1, padx=5, pady=5, sticky="ew")

    def create_sheet_selection_widgets(self):
        self.sheet_menu = ttk.OptionMenu(self.master, self.sheet_var, ())
        self.sheet_menu.config(width=20, bootstyle="outline-primary")
        self.sheet_menu.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

    def create_header_selection_widgets(self):
        self.header_frame = ttk.Frame(self.master)
        self.header_frame.grid(row=1, column=1, padx=5, pady=5, sticky="nsew")
        self.canvas = Canvas(self.header_frame)
        self.scrollbar = Scrollbar(self.header_frame, orient=VERTICAL, command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)
        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        ttk.Label(self.master, textvariable=self.selected_headers_var, bootstyle="info").grid(row=2, column=0, columnspan=2, padx=5, pady=5, sticky="ew")

    def create_content_header_widgets(self):
        ttk.Label(self.master, text="原文本内容", bootstyle="info").grid(row=3, column=0, padx=5, pady=5, sticky="ew")
        self.content_headers_menu = ttk.OptionMenu(self.master, self.content_header_var, ())
        self.content_headers_menu.config(width=20, bootstyle="outline-primary")
        self.content_headers_menu.grid(row=3, column=1, padx=5, pady=5, sticky="ew")

    def create_folder_selection_widgets(self):
        ttk.Button(self.master, text="选择音频文件夹", command=self.select_folder, bootstyle="outline-primary").grid(row=4, column=0, padx=5, pady=5, sticky="ew")
        ttk.Label(self.master, textvariable=self.folder_path_var, bootstyle="info").grid(row=4, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(self.master, text="选择结果保存文件夹", command=self.select_save_folder, bootstyle="outline-primary").grid(row=5, column=0, padx=5, pady=5, sticky="ew")
        ttk.Label(self.master, textvariable=self.save_folder_path_var, bootstyle="info").grid(row=5, column=1, padx=5, pady=5, sticky="ew")

    def create_server_widgets(self):
        ttk.Label(self.master, text="服务器地址:", bootstyle="info").grid(row=6, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(self.master, textvariable=self.server_url_var, bootstyle="primary").grid(row=6, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(self.master, text="检测服务器连接", command=self.check_server_connection, bootstyle="outline-primary").grid(row=7, column=0, columnspan=2, padx=5, pady=5, sticky="ew")

    def create_action_buttons(self):
        ttk.Button(self.master, text="音频检测", command=self.match_files, bootstyle="outline-primary").grid(row=8, column=0, columnspan=2, padx=5, pady=5, sticky="ew")

    def set_initial_window_size(self):
        self.master.update_idletasks()
        initial_width = self.master.winfo_reqwidth() + 20
        initial_height = self.master.winfo_reqheight() + 20
        self.master.geometry(f'{initial_width}x{initial_height}')

    def center_window(self):
        self.master.update_idletasks()
        width = self.master.winfo_width()
        height = self.master.winfo_height()
        x = (self.master.winfo_screenwidth() // 2) - (width // 2)
        y = (self.master.winfo_screenheight() // 2) - (height // 2)
        self.master.geometry(f'{width}x{height}+{x}+{y}')

    def select_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.file_path_var.set(file_path)
            self.load_sheets(file_path)

    def select_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.folder_path_var.set(folder_path)
            self.save_selection()

    def select_save_folder(self):
        save_folder_path = filedialog.askdirectory()
        if save_folder_path:
            self.save_folder_path_var.set(save_folder_path)
            self.save_selection()

    def load_sheets(self, file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet_names = workbook.sheetnames
        self.sheet_var.set(sheet_names[0])
        self.sheet_menu['menu'].delete(0, 'end')
        for sheet in sheet_names:
            self.sheet_menu['menu'].add_command(label=sheet, command=lambda value=sheet: self.sheet_var.set(value))
        self.sheet_var.trace_add('write', lambda *args: self.clear_and_load_headers(file_path))

    def clear_and_load_headers(self, file_path):
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        self.selected_headers = []
        self.selected_headers_var.set("音频文件名称格式: ")
        self.load_headers(file_path)

    def load_headers(self, file_path):
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[self.sheet_var.get()]
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1)) if cell.value]
        self.header_vars = {}
        self.content_headers_menu['menu'].delete(0, 'end')
        for header in headers:
            var = IntVar()
            chk = ttk.Checkbutton(self.scrollable_frame, text=header, variable=var, command=self.update_selected_headers, bootstyle="primary")
            chk.pack(anchor='w')
            self.header_vars[header] = var
            self.content_headers_menu['menu'].add_command(label=header, command=lambda value=header: self.content_header_var.set(value))

    def update_selected_headers(self):
        self.selected_headers = [header for header, var in self.header_vars.items() if var.get() == 1]
        self.selected_headers_var.set(f"音频文件名称格式: " + "_".join(self.selected_headers))
        self.save_selection()

    def load_previous_selection(self):
        try:
            with open("selected_headers.txt", "r") as f:
                data = f.read().splitlines()
                if data:
                    self.file_path_var.set(data[0])
                    self.load_sheets(data[0])
                    self.load_headers(data[0])
                    self.sheet_var.set(data[1])
                    self.selected_headers = data[2:-3]
                    self.folder_path_var.set(data[-3])
                    self.save_folder_path_var.set(data[-2])
                    self.content_header_var.set(data[-1])
                    for header in self.selected_headers:
                        if header in self.header_vars:
                            self.header_vars[header].set(1)
                    self.update_selected_headers()
        except Exception:
            pass

    def save_selection(self):
        with open("selected_headers.txt", "w") as f:
            f.write(f"{self.file_path_var.get()}\n")
            f.write(f"{self.sheet_var.get()}\n")
            f.write("\n".join(self.selected_headers))
            f.write(f"\n{self.folder_path_var.get()}\n")
            f.write(f"{self.save_folder_path_var.get()}\n")
            f.write(f"{self.content_header_var.get()}\n")

    def save_content_header_selection(self, *args):
        self.save_selection()

    def match_files(self):
        folder_path = self.folder_path_var.get()
        if not folder_path or folder_path == "选择一个文件夹":
            print("请先选择一个文件夹")
            return
        file_path = self.file_path_var.get()
        if not file_path or file_path == "选择一个XLSX文件":
            print("请先选择一个XLSX文件")
            return
        sheet_name = self.sheet_var.get()
        if not sheet_name or sheet_name == "选择一个Sheet":
            print("请先选择一个Sheet")
            return
        save_folder_path = self.save_folder_path_var.get()
        if not save_folder_path or save_folder_path == "选择结果保存文件夹":
            print("请先选择结果保存文件夹")
            return
        server_url = self.server_url_var.get()
        if not server_url:
            print("请先输入服务器URL")
            return

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        string_list, original_texts = self.extract_strings_and_texts(sheet)
        file_names = os.listdir(folder_path)
        string_map = self.map_strings_to_files(string_list, file_names)
        self.process_files_and_save_results(string_list, string_map, original_texts, folder_path, server_url, save_folder_path)

    def extract_strings_and_texts(self, sheet):
        string_list = []
        original_texts = []
        text_column_index = next((idx for idx, cell in enumerate(sheet[1]) if self.content_header_var.get() in str(cell.value)), None)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            selected_values = [str(row[header_index]) for header_index, header in enumerate(sheet[1]) if header.value in self.selected_headers]
            string_list.append("_".join(selected_values))
            original_text = row[text_column_index] if text_column_index is not None else ""
            original_texts.append(original_text)
        return string_list, original_texts

    def map_strings_to_files(self, string_list, file_names):
        audio_extensions = {'.mp3', '.wav', '.flac', '.aac', '.ogg', '.wma'}
        string_map = {string: [] for string in string_list}
        for file_name in file_names:
            if not any(file_name.lower().endswith(ext) for ext in audio_extensions):
                continue
            for string in string_list:
                if string in file_name:
                    string_map[string].append(file_name)
        return string_map

    def process_files_and_save_results(self, string_list, string_map, original_texts, folder_path, server_url, save_folder_path):
        progress_window, progress_bar, progress_info = self.create_progress_window()
        output_workbook, output_sheet = self.create_output_workbook()
        total_files = len(string_list)
        for i, string in enumerate(string_list):
            if not string_map[string]:  # If no files matched, still add to the output
                output_sheet.append([string, "", original_texts[i], "", "0.00%"])
            else:
                for file_name in string_map[string]:
                    audio_text, overlap_rate_percentage = self.process_single_file(string, file_name, original_texts, folder_path, server_url, i)
                    output_sheet.append([string, file_name, original_texts[i], audio_text, f"{overlap_rate_percentage:.2f}%"])
            self.update_progress(progress_bar, progress_info, i, total_files)
        output_workbook.save(os.path.join(save_folder_path, "匹配结果.xlsx"))
        progress_window.destroy()

    def create_progress_window(self):
        progress_window = ttk.Toplevel(self.master)
        progress_window.title("处理进度")
        ttk.Label(progress_window, text="正在处理音频文件...").pack(pady=10)
        progress_bar = ttk.Progressbar(progress_window, length=300, mode='determinate')
        progress_bar.pack(pady=10)
        progress_info = ttk.Label(progress_window, text="")
        progress_info.pack(pady=10)
        progress_window.update()
        return progress_window, progress_bar, progress_info

    def create_output_workbook(self):
        output_workbook = openpyxl.Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = "匹配结果"
        output_sheet.append(["字符串", "文件名", "原文本", "录制文本", "重合率"])
        return output_workbook, output_sheet

    def process_single_file(self, string, file_name, original_texts, folder_path, server_url, index):
        if not file_name:
            return "", 0

        audio_text = process_audio_file(os.path.join(folder_path, file_name), server_url)
        original_text = ''.join(ch for ch in original_texts[index] if '\u4e00' <= ch <= '\u9fff')
        recorded_text = ''.join(filter(str.isalnum, audio_text))
        overlap_rate = self.calculate_overlap_rate(original_text, recorded_text)
        overlap_rate_percentage = overlap_rate * 100

        return audio_text, overlap_rate_percentage

    def calculate_overlap_rate(self, original_text, recorded_text):
        original_text_set = set(original_text)
        recorded_text_set = set(recorded_text)
        overlap_count = len(original_text_set & recorded_text_set)
        return overlap_count / max(len(original_text_set), len(recorded_text_set)) if original_text_set and recorded_text_set else 0

    def update_progress(self, progress_bar, progress_info, current, total):
        progress_bar['value'] = (current + 1) / total * 100
        progress_info.config(text=f"已处理 {current + 1} 个文件，还剩 {total - (current + 1)} 个文件")
        progress_info.master.update()

    def check_server_connection(self):
        def show_result(success):
            result_window = ttk.Toplevel(self.master)
            result_window.title("服务器连接结果")
            ttk.Label(result_window, text="服务器连接成功" if success else "无法连接到服务器").pack(pady=10)
            ttk.Button(result_window, text="确定", command=result_window.destroy, bootstyle="outline-primary").pack(pady=10)
            self.center_popup(result_window)

        def async_check():
            success = check_server_connection(self.server_url_var.get())
            self.master.after(0, lambda: show_result(success))

        threading.Thread(target=async_check).start()

    def center_popup(self, window):
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)
        window.geometry(f'{width}x{height}+{x}+{y}')

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

if __name__ == "__main__":
    root = Tk()
    app = Application(root)
    root.mainloop()